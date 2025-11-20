#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
简易 Excel 表格检查脚本
用于读取《学生.xlsx》和《队长.xlsx》的表头与示例数据行，便于后续编写数据库初始化脚本。

依赖：openpyxl
运行：python scripts/inspect_excel.py
"""

import os
from typing import List, Tuple

try:
    from openpyxl import load_workbook
except Exception as e:
    raise SystemExit("请先安装 openpyxl，命令：pip3 install openpyxl")


def _row_to_str_list(row) -> List[str]:
    vals = []
    for v in row:
        if v is None:
            vals.append("")
        elif isinstance(v, str):
            vals.append(v.strip())
        else:
            vals.append(str(v))
    return vals


def detect_headers(ws, max_scan_rows: int = 10) -> Tuple[int, List[str]]:
    """在前 max_scan_rows 行中寻找表头行并返回其行号与表头列表"""
    candidate = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        values = _row_to_str_list(row)
        non_empty = [v for v in values if v]
        if not non_empty:
            continue
        # 简单启发式：非空且绝大多数是字符串，视为表头
        str_count = sum(1 for v in values if v and not any(ch.isdigit() for ch in v))
        if len(non_empty) >= max(2, int(len(values) * 0.5)) and str_count >= max(1, int(len(non_empty) * 0.6)):
            return idx, values
        # 兜底：记录第一个非空行
        if candidate is None:
            candidate = (idx, values)
    return candidate if candidate else (1, _row_to_str_list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))))


def inspect_excel(file_path: str, sample_rows: int = 5) -> None:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    sheet_name = ws.title
    header_row_idx, headers = detect_headers(ws)

    data_rows = []
    for r in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        vals = _row_to_str_list(r)
        if not any(vals):
            continue
        data_rows.append(vals)
        if len(data_rows) >= sample_rows:
            break

    print(f"文件: {file_path}")
    print(f"工作表: {sheet_name}")
    print(f"表头行号: {header_row_idx}")
    print("表头:")
    print(", ".join(headers))
    print(f"示例数据行(最多 {sample_rows} 行):")
    for i, row in enumerate(data_rows, start=1):
        print(f"{i}: " + ", ".join(row))
    print("-" * 60)


def main():
    base = os.getcwd()
    files = [
        os.path.join(base, "学生.xlsx"),
        os.path.join(base, "队长.xlsx"),
    ]
    for fp in files:
        if os.path.exists(fp):
            try:
                inspect_excel(fp)
            except Exception as e:
                print(f"[ERROR] 读取 {fp} 失败: {e}")
        else:
            print(f"[WARN] 未找到文件: {fp}")


if __name__ == "__main__":
    main()