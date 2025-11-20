#!/usr/bin/env python3
"""
数据库初始化脚本
支持：
1) 直接创建示例数据（无 Excel 文件时的兜底）
2) 从《学生.xlsx》《队长.xlsx》导入班级、学生、队长账号

Excel 期望格式：
《学生.xlsx》工作表名：信息表
表头：学号, 姓名, 警号, 年级, 区队, 区队(全称)
《队长.xlsx》表头：姓名, 职务（用户名将自动生成为 leader001…）
"""

from datetime import datetime
import os
import sys
from typing import List, Dict, Tuple
from werkzeug.security import generate_password_hash

## 保障包导入：将 src 目录加入搜索路径（scripts 的上一级）
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app import create_app, db
from app.models import User, Class, Achievement, AuditLog
from app.models.system_config import SystemConfig
try:
    from openpyxl import load_workbook
except ImportError:
    print("缺少 openpyxl 依赖，请先安装：pip3 install openpyxl")
    raise

COLLEGE_DEFAULT = '信息网络安全学院'
MAJOR_DEFAULT = None  # 无专业字段时置空


def load_students_from_excel(file_path: str) -> Tuple[List[Dict], Dict[str, Dict]]:
    """读取学生 Excel，返回学生列表和班级字典（按全称聚合）
    学生项：{'username','name','grade','team_short','team_full'}
    班级项：{'class_name','college','grade','major'}
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    # 表头定位
    headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # 简单映射
    try:
        idx_sid = headers.index('学号')
        idx_name = headers.index('姓名')
        idx_grade = headers.index('年级')
        # 两个“区队”列：短称与全称
        idx_team_short = headers.index('区队')
        # 找到第二个“区队”作为全称
        idx_team_full = headers.index('区队', idx_team_short + 1)
    except ValueError:
        raise RuntimeError('《学生.xlsx》表头不符合预期，请确保包含：学号, 姓名, 年级, 区队(两列)')

    students: List[Dict] = []
    classes_map: Dict[str, Dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[idx_sid]).strip() if row[idx_sid] is not None else ''
        name = str(row[idx_name]).strip() if row[idx_name] is not None else ''
        grade_raw = str(row[idx_grade]).strip() if row[idx_grade] is not None else ''
        team_short = str(row[idx_team_short]).strip() if row[idx_team_short] is not None else ''
        team_full = str(row[idx_team_full]).strip() if row[idx_team_full] is not None else ''
        if not sid or not name:
            continue
        # 年级标准化：如 "22" -> "2022"
        grade = grade_raw
        if grade_raw and len(grade_raw) == 2 and grade_raw.isdigit():
            grade = '20' + grade_raw

        students.append({
            'username': sid,
            'name': name,
            'grade': grade,
            'team_short': team_short,
            'team_full': team_full or team_short,
        })

        class_name = team_full or team_short or '未分配区队'
        if class_name not in classes_map:
            classes_map[class_name] = {
                'class_name': class_name,
                'college': COLLEGE_DEFAULT,
                'grade': grade,
                'major': MAJOR_DEFAULT,
            }
    return students, classes_map


def load_leaders_from_excel(file_path: str) -> List[Dict]:
    """读取队长 Excel，返回队长列表（用户名自动生成为 leader001…）
    队长项：{'username','name','duty'}
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        idx_name = headers.index('姓名')
        idx_duty = headers.index('职务')
    except ValueError:
        raise RuntimeError('《队长.xlsx》表头不符合预期，请确保包含：姓名, 职务')

    leaders: List[Dict] = []
    seq = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[idx_name]).strip() if row[idx_name] is not None else ''
        duty = str(row[idx_duty]).strip() if row[idx_duty] is not None else ''
        if not name:
            continue
        username = f'leader{seq:03d}'
        seq += 1
        leaders.append({'username': username, 'name': name, 'duty': duty})
    return leaders

def main():
    """初始化数据库"""
    app = create_app('development')
    
    with app.app_context():
        # 删除所有表（谨慎使用）
        print("正在删除现有表...")
        db.drop_all()
        
        # 创建所有表
        print("正在创建数据库表...")
        db.create_all()
        
        # 创建初始数据
        print("正在创建初始数据...")
        create_initial_data()
        
        print("数据库初始化完成！")

def create_initial_data():
    """创建初始数据：优先从 Excel 导入，缺失时回退到示例数据"""

    students_excel = os.path.join(os.path.dirname(__file__), '..', '..', '..', '学生.xlsx')
    students_excel = os.path.abspath(students_excel)
    leaders_excel = os.path.join(os.path.dirname(__file__), '..', '..', '..', '队长.xlsx')
    leaders_excel = os.path.abspath(leaders_excel)

    classes = []
    leaders = []
    students = []

    if os.path.exists(students_excel) and os.path.exists(leaders_excel):
        print(f"检测到 Excel 文件，开始导入：\n学生 -> {students_excel}\n队长 -> {leaders_excel}")
        # 读取 Excel
        students_list, class_map = load_students_from_excel(students_excel)
        leaders_list = load_leaders_from_excel(leaders_excel)

        # 创建班级
        for class_name, class_data in class_map.items():
            class_obj = Class(**class_data)
            db.session.add(class_obj)
            classes.append(class_obj)
        db.session.flush()  # 获取 class_id

        # 预计算初始化密码哈希，避免重复计算耗时
        admin_pwd_hash = generate_password_hash('admin123')
        leader_pwd_hash = generate_password_hash('leader123')
        default_pwd = SystemConfig.get_config('default_password', 'student123')
        default_pwd_hash = generate_password_hash(default_pwd)

        # 创建管理员
        admin = User(username='admin', name='系统管理员', role='admin')
        admin.password_hash = admin_pwd_hash
        db.session.add(admin)

        # 创建队长（不绑定班级）
        for item in leaders_list:
            leader = User(username=item['username'], name=item['name'], role='team_leader')
            leader.password_hash = leader_pwd_hash
            db.session.add(leader)
            leaders.append(leader)
        db.session.flush()

        # 创建学生
        # 建立班级名到ID的映射
        name_to_id = {c.class_name: c.class_id for c in classes}
        for s in students_list:
            class_id = name_to_id.get(s['team_full']) or name_to_id.get(s['team_short'])
            student = User(
                username=s['username'],
                name=s['name'],
                role='student',
                class_id=class_id
            )
            student.password_hash = default_pwd_hash
            db.session.add(student)
            students.append(student)

        # 可选：不自动创建示例成果，初始化更为干净
        db.session.commit()

        print("正在初始化系统配置...")
        SystemConfig.init_default_configs()

        print(f"创建了 {len(classes)} 个班级（来自学生表）")
        print("创建了 1 个管理员账户: admin/admin123")
        print(f"创建了 {len(leaders)} 个队长账户: leader001…/leader123")
        print(f"创建了 {len(students)} 个学生账户: 学号/{default_pwd}")
        print("系统配置已初始化")
        return

    # ===== 无 Excel 文件时，使用原示例数据 =====
    print("未检测到 Excel 文件，使用示例数据初始化。")
    classes_data = [
        {'class_name': '2021级侦查学1班', 'college': '侦查学院', 'grade': '2021', 'major': '侦查学'},
        {'class_name': '2021级侦查学2班', 'college': '侦查学院', 'grade': '2021', 'major': '侦查学'},
        {'class_name': '2021级刑事科学技术1班', 'college': '侦查学院', 'grade': '2021', 'major': '刑事科学技术'},
        {'class_name': '2021级网络安全与执法1班', 'college': '信息网络安全学院', 'grade': '2021', 'major': '网络安全与执法'},
        {'class_name': '2022级侦查学1班', 'college': '侦查学院', 'grade': '2022', 'major': '侦查学'},
    ]

    for class_data in classes_data:
        class_obj = Class(**class_data)
        db.session.add(class_obj)
        classes.append(class_obj)
    db.session.flush()

    # 预计算示例数据密码哈希
    admin_pwd_hash = generate_password_hash('admin123')
    leader_pwd_hash = generate_password_hash('leader123')
    student_pwd_hash = generate_password_hash('student123')

    admin = User(username='admin', name='系统管理员', role='admin')
    admin.password_hash = admin_pwd_hash
    db.session.add(admin)

    leaders_data = [
        {'username': 'leader001', 'name': '张队长'},
        {'username': 'leader002', 'name': '李队长'},
        {'username': 'leader003', 'name': '王队长'},
    ]
    for leader_data in leaders_data:
        leader = User(username=leader_data['username'], name=leader_data['name'], role='team_leader')
        leader.password_hash = leader_pwd_hash
        db.session.add(leader)
        leaders.append(leader)
    db.session.flush()

    students_data = [
        {'username': '2021001', 'name': '张三', 'class_idx': 0},
        {'username': '2021002', 'name': '李四', 'class_idx': 0},
        {'username': '2021003', 'name': '王五', 'class_idx': 1},
        {'username': '2021004', 'name': '赵六', 'class_idx': 1},
        {'username': '2021005', 'name': '钱七', 'class_idx': 2},
        {'username': '2022001', 'name': '孙八', 'class_idx': 4},
    ]
    for student_data in students_data:
        student = User(
            username=student_data['username'],
            name=student_data['name'],
            role='student',
            class_id=classes[student_data['class_idx']].class_id
        )
        student.password_hash = student_pwd_hash
        db.session.add(student)
        students.append(student)

    # 示例成果：包括各种状态的示例
    achievements_data = [
        {
            'title': '全国大学生数学建模竞赛一等奖',
            'type': 'competition',
            'level': 'national',
            'award_date': '2023-10-15',
            'supervisor': '张教授',
            'members': ['张三', '李四'],
            'submitter_idx': 0,
            'leader_idx': 0,
            'class_idx': 0,
            'status': 'approved'
        },
        {
            'title': '基于深度学习的图像识别算法研究',
            'type': 'paper',
            'level': 'province',
            'award_date': '2023-09-20',
            'supervisor': '李教授',
            'members': ['王五'],
            'submitter_idx': 2,
            'leader_idx': 1,
            'class_idx': 1,
            'status': 'pending'
        },
        {
            'title': '校级编程竞赛二等奖（草稿）',
            'type': 'competition',
            'level': 'school',
            'award_date': '2023-11-10',
            'supervisor': '王教授',
            'members': ['赵六', '孙七'],
            'submitter_idx': 1,
            'leader_idx': 0,
            'class_idx': 0,
            'status': 'draft'
        }
    ]
    for achievement_data in achievements_data:
        achievement = Achievement(
            title=achievement_data['title'],
            type=achievement_data['type'],
            level=achievement_data['level'],
            award_date=datetime.strptime(achievement_data['award_date'], '%Y-%m-%d').date(),
            supervisor=achievement_data['supervisor'],
            members=achievement_data['members'],
            submitter_id=students[achievement_data['submitter_idx']].user_id,
            leader_id=leaders[achievement_data['leader_idx']].user_id,
            class_id=classes[achievement_data['class_idx']].class_id,
            status=achievement_data['status']
        )
        db.session.add(achievement)

    db.session.commit()

    print("正在初始化系统配置...")
    SystemConfig.init_default_configs()

    print(f"创建了 {len(classes)} 个班级")
    print("创建了 1 个管理员账户: admin/admin123")
    print(f"创建了 {len(leaders)} 个队长账户: leader001-003/leader123")
    print(f"创建了 {len(students)} 个学生账户: 学号/student123")
    print(f"创建了 {len(achievements_data)} 个示例成果")
    print("系统配置已初始化")

if __name__ == '__main__':
    main()